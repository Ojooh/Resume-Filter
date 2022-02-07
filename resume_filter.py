import  openpyxl, smtplib, sys, imaplib, email
import pprint, os, re, PyPDF2, docx, datetime



class ResumeFilter:
    def __init__(self, email, password, subject=["application"], skills=None, server="mail.zkyte.com.ng", reply='N'):
        self.pat            = re.compile(r'\s+')
        self.pat1           = re.compile(r'\n+')
        self.email          = email
        self.passy          = password
        self.subject        = subject
        self.server         = server
        self.skillset       = skills
        self.reply          = reply
        self.from_data      = {}
        

    def setUpDirectory(self):
        detach_dir          = 'resumes-and-candidate-data\\'
        now_time            = str(datetime.datetime.now())
        micro_second_index  = now_time.index('.')
        now_time            = now_time[:micro_second_index]
        detach_dir          = detach_dir+now_time
        detach_dir          = detach_dir.replace(' ',',')
        self.detach_dir     = detach_dir.replace(':','-')

        if not os.path.exists(self.detach_dir):
        	os.makedirs(self.detach_dir)

        return True


    def getSearchQuery(self):
        query = ''
        for g in self.subject:
            query += '( SUBJECT "' + g + '") '

        return query
            


    def inboxSearch(self):
        print('Searching for the resumes...\n\n')
        state = False
        file_type = ""

        try:
            m = imaplib.IMAP4_SSL(self.server)
            m.login(self.email, self.passy)
        except:
            m = None

        if m is not None:
            m.select("inbox")
            # # '(OR (TO "tech163@fusionswift.com") (FROM "tech163@fusionswift.com"))'
            resp, items     = m.search(None,'(OR ' + self.getSearchQuery() + ')')
            items           = items[0].split()
            print(items)

            # while(len(items) > 0):
            for i in range(4):
                try:
                    emailid     = items[len(items) - 1]
                    resp, data  = m.fetch(emailid, "(RFC822)")
                    email_body  = data[0][1]
                    email_body  = email_body.decode('utf-8')
                    mail        = email.message_from_string(email_body)
                    # temp        = m.store(emailid,'+FLAGS', '\\Seen')
                    # m.expunge()
                    removed     = items.pop()

                    if mail.get_content_maintype() != 'multipart':
                        continue

                    received_from           = mail["From"]
                    email_start_index       = received_from.index('<') + 1
                    email_end_index         = received_from.index('>')
                    received_from_emailid   = received_from[email_start_index:email_end_index]
                    received_from_name      = received_from[:email_start_index - 1]
                    received_from_date      = mail["Date"]

                    print ("["+mail["From"]+"] :" + mail["Subject"])

                    for part in mail.walk():
                        if part.get_filename() is not None:
                            file_type   = "." + part.get_filename().split(".")[-1]
                        if part.get_content_maintype() == 'multipart':
                            continue
                        if part.get('Content-Disposition') is None:
                            continue
                        if part.get_filename().endswith('.pdf'):
                            file_type = '.pdf'
                        if part.get_filename().endswith('.docx'):
                            file_type='.docx'
                    
                        filename    = received_from_emailid + file_type
                        att_path    = os.path.join(self.detach_dir, filename)

                        if not os.path.isfile(att_path) :
                            fp = open(att_path, 'wb')
                            fp.write(part.get_payload(decode=True))
                            fp.close()
                            self.from_data[received_from_emailid]   = [received_from_emailid, received_from_name, received_from_date, att_path]
                    state = True
                except Exception as e:
                    asdf=1 #do nothing
                    state = False
                    print("something went wrong " + str(e))

            print ('Finished downloading resumes.\n\n')
        else:
            print("Could not connect to email Account, plese check credentials and try again.")

        return state


    def reset_eof_of_pdf_return_stream(self, pdf_stream_in:list):
        for i, x in enumerate(pdf_stream_in[::-1]):
            if b'%%EOF' in x:
                actual_line = len(pdf_stream_in)-i
                print(f'EOF found at line position {-i} = actual {actual_line}, with value {x}')
                break

        # return the list up to that point
        return pdf_stream_in[:actual_line]


    def confirmEOF(self, path):
        with open(path, 'rb') as p:
            txt = (p.readlines())

        txtx = self.reset_eof_of_pdf_return_stream(txt)

        with open(path, 'wb') as f:
            f.writelines(txtx)


    def extractText(self):
        print('Scanning all the resumes...\n\n')
        pat         = re.compile(r'\s+')
        pat1        = re.compile(r'\n+')
        pat2        = re.compile(r'\n+')
        phoney      = re.compile(r'(\d{3}[-\.\s]??\d{3}[-\.\s]??\d{4}|\(\d{3}\)\s*\d{3}[-\.\s]??\d{4}|\d{3}[-\.\s]??\d{4})')
        regex       = r"(?i)\b((?:https?://|www\d{0,3}[.]|[a-z0-9.\-]+[.][a-z]{2,4}/)(?:[^\s()<>]+|\(([^\s()<>]+|(\([^\s()<>]+\)))*\))+(?:\(([^\s()<>]+|(\([^\s()<>]+\)))*\)|[^\s`!()\[\]{};:'\".,<>?«»“”‘’]))"
        pat3        = re.compile(r'\d{3}[-,\s]\d{3}[-,\s]\d{4}')

        for downloaded_resume in self.from_data:
            content=''
            weights     = 0
            total_score = 0
            avg         =   0

            if self.from_data[downloaded_resume][3].endswith('.pdf'):
                path        = self.from_data[downloaded_resume][3]
                self.confirmEOF(path)
                pdfFileObj  = open(path, 'rb')
                pdfReader   = PyPDF2.PdfFileReader(pdfFileObj)
                content     = ''
                pages       = pdfReader.numPages

                for i in range(pages):
                    pageObj     = pdfReader.getPage(i)
                    content     += pageObj.extractText()
                content     = pat.sub('',content)
                content     = pat1.sub('',content)
                


            if self.from_data[downloaded_resume][3].endswith('.docx'):
                content = []
                path    = self.from_data[downloaded_resume][3]
                doc     = docx.Document(path)

                for para in doc.paragraphs:
                    content.append(para.text)

                content = ''.join(content)
                # print(content)
                content = pat.sub('',content)
                content = pat1.sub('',content)
                # print(content)

            content             = content.lower()
            phones_numbers1     = pat2.findall(content)
            # phones_numbers2     = pat3.findall(content)
            phones_numbers2     = phoney.findall(content)
            urls                = re.findall(regex, content)
            urls_all            = ', \n'.join(urls)
            phones_numbers_all  = ','.join(phones_numbers1) +  ','.join(phones_numbers2)
		    # phones_numbers_all  = ','.join(phones_numbers1) + ','.join(phones_numbers2)

            if len(phones_numbers_all) < 2:
                self.from_data[downloaded_resume].append('-')
    			# self.from_data[downloaded_resume].append('-')
            else:
                self.from_data[downloaded_resume].append(phones_numbers_all)

            if len(urls) > 0:
                self.from_data[downloaded_resume].append(urls_all)
            else :
                self.from_data[downloaded_resume].append('')

            
            if self.skillset is not None and len(self.skillset) > 0:
                # print(content)
                for required_skill in self.skillset:
                    total_score  += float(required_skill["score"])
                    if required_skill["skill"].lower() in content.lower():
                        print(str(self.from_data[downloaded_resume]) + " Applicant has this skill" + str(required_skill["skill"]))
                        weights += float(required_skill["score"])

                weights = weights / len(self.skillset)
                avg = total_score / len(self.skillset)

            geights = str(weights) + "/ " + str(avg)
            self.from_data[downloaded_resume].append(geights)
            if weights >  avg:
                self.from_data[downloaded_resume].append('yes')
            elif weights == avg:
                self.from_data[downloaded_resume].append('No decision made')
            else:
                self.from_data[downloaded_resume].append('No')

        print('Finished scanning all the resumes.\n\n')


    def saveInXl(self):
        print(self.from_data)
    	#self.from_data[id]=[id,name,date,filepath,number,links,score, decision]
        print('Saving data in excel sheet...\n')

        wb          = openpyxl.Workbook()
        sheet       = wb.active
        sheet.ttile = 'resume filter result'
        sheet.cell(row=1,column=1).value    = 'NAME'
        sheet.cell(row=1,column=2).value    = 'PHONE NUMBER'
        sheet.cell(row=1,column=3).value    = 'EMAIL ID'
        sheet.cell(row=1,column=4).value    = 'LINKS'
        sheet.cell(row=1,column=5).value    = 'DATE-TIME'
        sheet.cell(row=1,column=6).value    = 'DECISION'
        sheet.cell(row=1,column=7).value    = 'SCORE'
        sheet_row                           = 2

        for downloaded_resume in self.from_data:
            sheet.cell(row=sheet_row,column=1).value    = self.from_data[downloaded_resume][1]
            sheet.cell(row=sheet_row,column=2).value    = self.from_data[downloaded_resume][4]
            sheet.cell(row=sheet_row,column=3).value    = self.from_data[downloaded_resume][0]
            sheet.cell(row=sheet_row,column=4).value    = self.from_data[downloaded_resume][5]
            sheet.cell(row=sheet_row,column=5).value    = self.from_data[downloaded_resume][2]
            sheet.cell(row=sheet_row,column=6).value    = self.from_data[downloaded_resume][7]
            sheet.cell(row=sheet_row,column=7).value    = self.from_data[downloaded_resume][6]
            sheet_row                                   += 1

        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['C'].width = 40
        sheet.column_dimensions['D'].width = 40
        sheet.column_dimensions['E'].width = 40
        sheet.column_dimensions['F'].width = 40
        sheet.column_dimensions['G'].width = 20
        wb.save(self.detach_dir+ '\\candidate_data.xlsx')
        print("Finished saving data in excel sheet.\n\n")


    def sendMail(self):
        print("Sending replies to candidates...\n ")
        try:
            smtpObj     = smtplib.SMTP('smtp.zkyte.com.ng', 587)
            smtpObj.ehlo()
            smtpObj.starttls()
            smtpObj.login(self.email, self.passy)
        except Exception as e:
            msg = str(e)
            smtpObj = None


        if smtpObj is not None:
            wb          = openpyxl.load_workbook(self.detach_dir+ '\\candidate_data.xlsx')
            sheet       = wb.get_sheet_by_name('resumes')
            lastCol     = 6

            for r in range(2, 2 + len(self.from_data)):
                decision    = sheet.cell(row=r, column=lastCol).value
                name        = sheet.cell(row=r, column=1).value
                senderemail = sheet.cell(row=r, column=3).value

                if decision == 'Yes':
                    body = "Subject: SELECTED.\nDear %s,\n We are glad to inform you that you are selected for next phase of interview." %(name)
                else:
                    body = "Subject: rejected.\nDear %s,\n We are sorry to inform you that you are not selected for techincal interview." %(name)

                print('Sending email to %s...' % senderemail)
                # sendmailStatus = smtpObj.sendmail(useremail, senderemail, body)

                if sendmailStatus != {}:
                    print('There was a problem sending email to %s: %s' % (sendereemail,sendmailStatus))

                smtpObj.quit()
                print("Finished sending replies to candidates.\n\n")
        
        else:
            print('There was a problem sending email to' + str(msg))


        
		
		

    def filterResume(self):
        self.setUpDirectory()
        self.inboxSearch()
        self.extractText()
        self.saveInXl() 

        if self.reply.lower() == 'y':
            self.sendMail() 
        
        
        
        
        
        

        

        
            


            

	        
	        

